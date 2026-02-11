import os
import sys
import cv2
import numpy as np
import pandas as pd
from PIL import Image
from matplotlib import pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from scipy import stats
import warnings
from datetime import datetime

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QGridLayout, QPushButton, QLabel, QFrame, QSplitter, QComboBox, 
    QLineEdit, QRadioButton, QSpinBox, QCheckBox, QGroupBox, 
    QGraphicsView, QGraphicsScene, QGraphicsRectItem, QGraphicsLineItem, 
    QTreeWidget, QTreeWidgetItem, QMessageBox, QFileDialog, QSlider, 
    QDialog, QTabWidget, QFormLayout, QButtonGroup, QMenu
)
from PySide6.QtCore import Qt, QRectF, QPointF, Signal, Slot, QMimeData
from PySide6.QtGui import QPixmap, QImage, QPen, QColor, QPainter, QIcon, QFont

class ProfileDialog(QDialog):
    def __init__(self, roi_data, separators, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Lane Intensity Profile")
        self.resize(600, 400)
        
        layout = QVBoxLayout(self)
        
        # Create figure and canvas
        self.figure = Figure(figsize=(6, 4), dpi=100)
        self.canvas = FigureCanvas(self.figure)
        layout.addWidget(self.canvas)
        
        self.plot_profile(roi_data, separators)
        
    def plot_profile(self, roi_data, separators):
        # Invert to make bands peaks
        roi_data = cv2.bitwise_not(roi_data)
        
        # Calculate vertical profile (mean across width)
        # Note: We want to see peaks across the width of the ROI
        # so we average along the height (axis=0)
        profile = np.mean(roi_data, axis=0)
        
        ax = self.figure.add_subplot(111)
        ax.plot(profile, color='#3498db', linewidth=2)
        ax.set_title("Intensity Profile (Across ROI Width)")
        ax.set_xlabel("Pixel Position (X)")
        ax.set_ylabel("Mean Intensity")
        
        # Add separator lines
        for sep in separators:
            ax.axvline(x=sep, color='#e74c3c', linestyle='--', alpha=0.7)
            
        ax.grid(True, alpha=0.3)
        self.figure.tight_layout()
        self.canvas.draw()

class AnalysisCanvas(QGraphicsView):
    roi_selected = Signal(QRectF)
    separator_moved = Signal(int, float) # index, new_x
    
    def __init__(self, scene, parent=None):
        super().__init__(scene, parent)
        self.setRenderHint(QPainter.Antialiasing)
        self.setDragMode(QGraphicsView.NoDrag)
        
        self.current_rect_item = None
        self.start_pos = None
        self.is_drawing = False
        self.is_resizing = False
        self.is_dragging_sep = False
        self.active_sep = None
        self.resize_edge = None
        self.separators = []
        
        # Pen styles
        self.roi_pen = QPen(QColor("#3498db"), 2)
        self.roi_pen.setStyle(Qt.DashLine)
        self.sep_pen = QPen(QColor("#e74c3c"), 2)

    def mousePressEvent(self, event):
        pos = self.mapToScene(event.pos())
        
        # Check if clicking on a separator for dragging
        for i, sep in enumerate(self.separators):
            if abs(sep.line().p1().x() - pos.x()) < 5:
                self.is_dragging_sep = True
                self.active_sep = i
                return

        # Check for resizing existing rect
        if self.current_rect_item:
            rect = self.current_rect_item.rect()
            edge_threshold = 10
            
            near_left = abs(pos.x() - rect.left()) < edge_threshold
            near_right = abs(pos.x() - rect.right()) < edge_threshold
            near_top = abs(pos.y() - rect.top()) < edge_threshold
            near_bottom = abs(pos.y() - rect.bottom()) < edge_threshold
            
            if (near_left or near_right) and rect.top() <= pos.y() <= rect.bottom():
                self.is_resizing = True
                self.resize_edge = 'left' if near_left else 'right'
                return
            elif (near_top or near_bottom) and rect.left() <= pos.x() <= rect.right():
                self.is_resizing = True
                self.resize_edge = 'top' if near_top else 'bottom'
                return
            
            # If click is INSIDE the box but not near edges or separators, ignore (don't start new drawing)
            if rect.contains(pos):
                return

        # Start new drawing
        self.is_drawing = True
        self.start_pos = pos
        self.clear_selection()
            
        self.current_rect_item = QGraphicsRectItem(QRectF(pos, pos))
        self.current_rect_item.setPen(self.roi_pen)
        self.scene().addItem(self.current_rect_item)

    def mouseMoveEvent(self, event):
        pos = self.mapToScene(event.pos())
        
        # Priority: Dragging separator
        if self.is_dragging_sep:
            if not self.current_rect_item:
                self.is_dragging_sep = False
                return
            rect = self.current_rect_item.rect()
            new_x = max(rect.left(), min(rect.right(), pos.x()))
            line = self.separators[self.active_sep].line()
            line.setP1(QPointF(new_x, rect.top()))
            line.setP2(QPointF(new_x, rect.bottom()))
            self.separators[self.active_sep].setLine(line)
            self.setCursor(Qt.ClosedHandCursor) # Visual feedback during drag
            return

        # Check for separator hover
        found_sep = False
        for sep in self.separators:
            try:
                if abs(sep.line().p1().x() - pos.x()) < 8:
                    y_min = min(sep.line().p1().y(), sep.line().p2().y())
                    y_max = max(sep.line().p1().y(), sep.line().p2().y())
                    if y_min - 5 <= pos.y() <= y_max + 5:
                        self.setCursor(Qt.PointingHandCursor)
                        found_sep = True
                        break
            except RuntimeError: continue
        
        if found_sep: return

        if not self.is_drawing and not self.is_resizing:
            if self.current_rect_item:
                try:
                    rect = self.current_rect_item.rect()
                    margin = 10
                    on_left = abs(pos.x() - rect.left()) < margin
                    on_right = abs(pos.x() - rect.right()) < margin
                    on_top = abs(pos.y() - rect.top()) < margin
                    on_bottom = abs(pos.y() - rect.bottom()) < margin

                    if (on_left or on_right) and rect.top() <= pos.y() <= rect.bottom():
                        self.setCursor(Qt.SizeHorCursor)
                    elif (on_top or on_bottom) and rect.left() <= pos.x() <= rect.right():
                        self.setCursor(Qt.SizeVerCursor)
                    else:
                        self.setCursor(Qt.ArrowCursor)
                except RuntimeError:
                    self.current_rect_item = None
                    self.setCursor(Qt.ArrowCursor)
            else:
                self.setCursor(Qt.ArrowCursor)
            return

        if self.is_drawing:
            if not self.current_rect_item:
                return
            rect = QRectF(self.start_pos, pos).normalized()
            self.current_rect_item.setRect(rect)
            self.roi_selected.emit(rect)
        elif self.is_resizing:
            if not self.current_rect_item:
                self.is_resizing = False
                return
            rect = self.current_rect_item.rect()
            if self.resize_edge == 'left':
                rect.setLeft(pos.x())
            elif self.resize_edge == 'right':
                rect.setRight(pos.x())
            elif self.resize_edge == 'top':
                rect.setTop(pos.y())
            elif self.resize_edge == 'bottom':
                rect.setBottom(pos.y())
            new_rect = rect.normalized()
            self.current_rect_item.setRect(new_rect)
            self.roi_selected.emit(new_rect)
        else:
            self.setCursor(Qt.ArrowCursor)

    def mouseReleaseEvent(self, event):
        if self.is_dragging_sep:
            self.setCursor(Qt.ArrowCursor)
            
        self.is_drawing = False
        self.is_resizing = False
        self.is_dragging_sep = False
        self.active_sep = None
        self.resize_edge = None
        if self.current_rect_item:
            try:
                self.roi_selected.emit(self.current_rect_item.rect())
            except RuntimeError:
                self.current_rect_item = None

    def update_separators(self):
        # The main logic for creating separators remains in BlotQuant
        # This method is just a placeholder to maintain consistency
        pass

    def clear_selection(self):
        if self.current_rect_item:
            try:
                self.scene().removeItem(self.current_rect_item)
            except (RuntimeError, AttributeError):
                pass
            self.current_rect_item = None
        
        for sep in self.separators:
            try:
                self.scene().removeItem(sep)
            except (RuntimeError, AttributeError):
                pass
        self.separators = []

class BlotQuant(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("BlotQuant v1.1.0")
        self.resize(1600, 800)
        
        # Set Window Icon
        icon_path = "Blot.ico"
        if hasattr(sys, '_MEIPASS'):
            icon_path = os.path.join(sys._MEIPASS, "Blot.ico")
        
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        
        self.version = "1.1.2"
        self.creation_date = "2026-02-11"
        self.author = "Dr. Robert Hauffe"
        self.affiliation = "Molecular and Experimental Nutritional Medicine, University of Potsdam, Germany"
        
        # Core State
        self.image = None
        self.display_width = 800
        self.display_height = 600
        self.rotation_angle = 0
        self.current_roi_rect = None
        self.separators = []
        self.excluded_samples = {} # {group_name: [list of excluded indices]}
        self.group_history = [] # For dropdown selection
        
        # Analysis Data
        self.analysis_history = [] # List of {'type': 'Loading Control'|'Target', 'group': str, 'detail': str, 'name': str, 'intensities': list}
        
        self.setup_ui()
        
    def setup_ui(self):
        # Main Widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        # 1. Header
        header_layout = QHBoxLayout()
        title_label = QLabel("BlotQuant")
        title_label.setStyleSheet("font-size: 28pt; font-weight: bold; color: #2c3e50;")
        header_layout.addWidget(title_label)
        
        header_layout.addStretch()
        
        about_btn = QPushButton("About BlotQuant")
        about_btn.setFixedSize(120, 30)
        about_btn.clicked.connect(self.show_about)
        header_layout.addWidget(about_btn)
        
        main_layout.addLayout(header_layout)
        
        # 2. Splitter for Left/Right
        splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(splitter)
        
        # Left Panel (Controls + View)
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        splitter.addWidget(left_panel)
        
        # Action Buttons
        btn_layout = QHBoxLayout()
        self.load_btn = self.create_button("Load Image", "#3498db", self.load_image)
        self.graphpad_btn = self.create_button("Open GraphPad", "#e67e22", self.open_graphpad)
        self.undo_btn = self.create_button("Undo Last", "#f1c40f", self.undo_last)
        self.reset_btn = self.create_button("Start Over", "#e74c3c", self.start_over)
        
        for btn in [self.load_btn, self.graphpad_btn, self.undo_btn, self.reset_btn]:
            btn_layout.addWidget(btn)
        left_layout.addLayout(btn_layout)
        
        # Management Area (Analysis Settings)
        mgmt_layout = QHBoxLayout()
        
        # Analysis Settings
        settings_group = QGroupBox("Analysis Settings")
        settings_group.setMinimumWidth(850)  # Widen to ensure everything fits
        settings_layout = QVBoxLayout(settings_group)
        settings_layout.setAlignment(Qt.AlignLeft)  # Left align internal layout
        
        # Add Experiment and Help Button to Group Box Header Area
        header_with_help = QHBoxLayout()
        header_with_help.setAlignment(Qt.AlignLeft)
        header_with_help.addWidget(QLabel("Experiment:"))
        self.experiment_input = QLineEdit()
        self.experiment_input.setPlaceholderText("e.g. Insulin Stimulation #1")
        self.experiment_input.setFixedWidth(300)  # Slightly wider
        header_with_help.addWidget(self.experiment_input)
        
        header_with_help.addSpacing(20)
        header_with_help.addStretch()
        
        help_btn = QPushButton("?")
        help_btn.setFixedSize(20, 20)
        help_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border-radius: 10px;
                font-weight: bold;
                font-size: 10pt;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        help_btn.setToolTip("Click for settings explanation")
        help_btn.clicked.connect(self.show_settings_help)
        header_with_help.addWidget(help_btn)
        settings_layout.addLayout(header_with_help)
        
        # Grid layout for Type and Group settings to ensure alignment
        grid_settings = QGridLayout()
        grid_settings.setAlignment(Qt.AlignLeft)
        grid_settings.setHorizontalSpacing(15)
        grid_settings.setColumnStretch(5, 1) # Add stretch to the end
        
        # Row 1: Mode (Type)
        grid_settings.addWidget(QLabel("Type:"), 0, 0)
        self.mode_group = QButtonGroup(self)
        self.mode_control = QRadioButton("Loading Control")
        self.mode_target = QRadioButton("Target")
        self.mode_control.setChecked(True)
        self.mode_group.addButton(self.mode_control)
        self.mode_group.addButton(self.mode_target)
        grid_settings.addWidget(self.mode_control, 0, 1)
        grid_settings.addWidget(self.mode_target, 0, 2)
        
        grid_settings.addWidget(QLabel("Protein Name:"), 0, 3)
        self.protein_name_input = QLineEdit()
        self.protein_name_input.setPlaceholderText("e.g. pAKT, Ponceau, etc.")
        self.protein_name_input.setFixedWidth(200)
        grid_settings.addWidget(self.protein_name_input, 0, 4)
        
        # Row 2: Group
        grid_settings.addWidget(QLabel("Group:"), 1, 0)
        self.analysis_group_group = QButtonGroup(self)
        self.group_ctrl = QRadioButton("Control")
        self.group_treat = QRadioButton("Treatment")
        self.group_ctrl.setChecked(True)
        self.analysis_group_group.addButton(self.group_ctrl)
        self.analysis_group_group.addButton(self.group_treat)
        grid_settings.addWidget(self.group_ctrl, 1, 1)
        grid_settings.addWidget(self.group_treat, 1, 2)
        
        grid_settings.addWidget(QLabel("Group Name:"), 1, 3)
        self.treatment_detail_input = QComboBox()
        self.treatment_detail_input.setEditable(True)
        self.treatment_detail_input.setPlaceholderText("e.g. Ctrl, Insulin 10nM, etc.")
        self.treatment_detail_input.lineEdit().setPlaceholderText("e.g. Ctrl, Insulin 10nM, etc.")
        self.treatment_detail_input.setFixedWidth(200)
        grid_settings.addWidget(self.treatment_detail_input, 1, 4)
        
        # Row 3: Replicates & Start Lane
        grid_settings.addWidget(QLabel("Replicates:"), 2, 0)
        self.reps_spin = QSpinBox()
        self.reps_spin.setRange(1, 20)
        self.reps_spin.setValue(6)
        self.reps_spin.setMinimumWidth(60) # Increased width for visibility
        grid_settings.addWidget(self.reps_spin, 2, 1)
        
        self.equal_n_check = QCheckBox("Equal N")
        grid_settings.addWidget(self.equal_n_check, 2, 2)
        
        grid_settings.addWidget(QLabel("Start Lane:"), 2, 3)
        self.start_idx_spin = QSpinBox()
        self.start_idx_spin.setRange(1, 20)
        self.start_idx_spin.setValue(1)
        self.start_idx_spin.setMinimumWidth(60) # Increased width for visibility
        grid_settings.addWidget(self.start_idx_spin, 2, 4)
        
        settings_layout.addLayout(grid_settings)
        
        # Options Row
        options_layout = QHBoxLayout()
        options_layout.setAlignment(Qt.AlignLeft)
        self.lock_roi_check = QCheckBox("Lock ROI Size")
        self.lock_roi_check.setToolTip("Forces new ROI to match the dimensions of the previous selection")
        options_layout.addWidget(self.lock_roi_check)
        
        options_layout.addSpacing(15)
        self.ttest_check = QCheckBox("Perform Welch's Test")
        self.ttest_check.setChecked(True)
        options_layout.addWidget(self.ttest_check)
        options_layout.addStretch()
        settings_layout.addLayout(options_layout)
        
        mgmt_layout.addWidget(settings_group)
        mgmt_layout.addStretch() # Push the group to the left
        left_layout.addLayout(mgmt_layout)
        
        # Rotation Control
        rotation_layout = QHBoxLayout()
        rotation_layout.addWidget(QLabel("Rotation Alignment:"))
        self.rotation_slider = QSlider(Qt.Horizontal)
        self.rotation_slider.setRange(-45, 45)
        self.rotation_slider.setValue(0)
        self.rotation_slider.valueChanged.connect(self.rotate_image)
        rotation_layout.addWidget(self.rotation_slider)
        reset_rot_btn = QPushButton("Reset")
        reset_rot_btn.clicked.connect(self.reset_rotation)
        rotation_layout.addWidget(reset_rot_btn)
        left_layout.addLayout(rotation_layout)
        
        # Image View Area
        view_container = QHBoxLayout()
        self.scene = QGraphicsScene()
        self.graphics_view = AnalysisCanvas(self.scene)
        self.graphics_view.roi_selected.connect(self.on_roi_selected)
        view_container.addWidget(self.graphics_view, stretch=4)
        
        apply_layout = QVBoxLayout()
        apply_layout.addStretch()
        
        apply_btn = QPushButton("QUANTIFY\nSELECTION")
        apply_btn.setFixedSize(110, 80)
        apply_btn.setStyleSheet("background-color: #2ecc71; color: white; font-weight: bold; border-radius: 5px;")
        apply_btn.clicked.connect(self.apply_selection)
        apply_layout.addWidget(apply_btn)

        apply_layout.addSpacing(10)

        profile_btn = QPushButton("SHOW\nPROFILE")
        profile_btn.setFixedSize(110, 60)
        profile_btn.setStyleSheet("background-color: #3498db; color: white; font-weight: bold; border-radius: 5px; font-size: 8pt;")
        profile_btn.clicked.connect(self.show_profile)
        apply_layout.addWidget(profile_btn)
        
        apply_layout.addSpacing(10)
        
        self.copy_btn = QPushButton("COPY DATA\nTO CLIPBOARD")
        self.copy_btn.setFixedSize(110, 60)
        self.copy_btn.setStyleSheet("background-color: #9b59b6; color: white; font-weight: bold; border-radius: 5px; font-size: 8pt;")
        self.copy_btn.clicked.connect(self.copy_to_clipboard)
        apply_layout.addWidget(self.copy_btn)
        
        apply_layout.addSpacing(10)
        
        self.export_btn = QPushButton("EXPORT\nTO EXCEL")
        self.export_btn.setFixedSize(110, 60)
        self.export_btn.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; border-radius: 5px; font-size: 8pt;")
        self.export_btn.clicked.connect(self.export_data)
        apply_layout.addWidget(self.export_btn)
        
        apply_layout.addStretch()
        view_container.addLayout(apply_layout)
        
        left_layout.addLayout(view_container)
        
        # Right Panel (Results)
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        splitter.addWidget(right_panel)
        
        results_header = QHBoxLayout()
        results_header.addWidget(QLabel("Analysis Summary"))
        results_header.addStretch()
        
        results_help_btn = QPushButton("?")
        results_help_btn.setFixedSize(20, 20)
        results_help_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border-radius: 10px;
                font-weight: bold;
                font-size: 10pt;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        results_help_btn.setToolTip("Click for results explanation")
        results_help_btn.clicked.connect(self.show_results_help)
        results_header.addWidget(results_help_btn)
        
        right_layout.addLayout(results_header)
        
        self.summary_text = QLabel("No analysis performed yet.")
        self.summary_text.setStyleSheet("background-color: #f8f9fa; border: 1px solid #dee2e6; padding: 10px; font-family: monospace;")
        self.summary_text.setWordWrap(True)
        right_layout.addWidget(self.summary_text)

        right_layout.addWidget(QLabel("Detailed Data Points"))
        self.results_tree = QTreeWidget()
        self.results_tree.setHeaderLabels(["Sample ID", "Loading Control", "Target", "Normalized"])
        self.results_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.results_tree.customContextMenuRequested.connect(self.show_tree_context_menu)
        self.results_tree.setAlternatingRowColors(True)
        self.results_tree.header().setStretchLastSection(False)
        self.results_tree.setHorizontalScrollMode(QTreeWidget.ScrollPerPixel)
        self.results_tree.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        right_layout.addWidget(self.results_tree)
        
        # Proportions: Left: 1, Right: 1
        # Balanced view between image and results
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 1)

    def create_button(self, text, color, slot):
        btn = QPushButton(text)
        btn.setStyleSheet(f"background-color: {color}; color: white; font-weight: bold; padding: 5px;")
        btn.clicked.connect(slot)
        return btn

    def show_tree_context_menu(self, pos):
        item = self.results_tree.itemAt(pos)
        if not item or item.parent() is None: # Only allow exclusion for replicates, not parent groups
            return
            
        menu = QMenu(self)
        group_name = item.parent().text(0)
        replicate_idx = item.parent().indexOfChild(item)
        
        is_excluded = replicate_idx in self.excluded_samples.get(group_name, [])
        
        action_text = "Include in Analysis" if is_excluded else "Exclude from Analysis"
        toggle_action = menu.addAction(action_text)
        toggle_action.triggered.connect(lambda: self.toggle_exclusion(group_name, replicate_idx))
        
        menu.exec(self.results_tree.mapToGlobal(pos))

    def toggle_exclusion(self, group_name, idx):
        if group_name not in self.excluded_samples:
            self.excluded_samples[group_name] = []
            
        if idx in self.excluded_samples[group_name]:
            self.excluded_samples[group_name].remove(idx)
        else:
            self.excluded_samples[group_name].append(idx)
            
        self.refresh_analysis()

    def show_settings_help(self):
        help_msg = (
            "<h3>Analysis Settings Guide</h3>"
            "<b>Type:</b> Choose <i>Loading Control</i> (e.g. Actin, Ponceau) for normalization, "
            "or <i>Target</i> for the protein of interest.<br><br>"
            "<b>Protein Name:</b> The label that will appear in the results and Excel export.<br><br>"
            "<b>Group:</b> Classify your selection as <i>Control</i> or <i>Treatment</i> for statistical comparison.<br><br>"
            "<b>Group Name:</b> A detailed description (e.g. 'Ctrl', 'Insulin 10nM'). This dropdown saves your entries for quick selection across multiple blots.<br><br>"
            "<b>Replicates:</b> The number of lanes/bands in your current ROI selection.<br><br>"
            "<b>Start Lane:</b> The starting index for lane numbering (e.g. if Lane 1-6 are on Blot A, set Blot B to start at #7).<br><br>"
            "<b>Equal N:</b> If checked, BlotQuant expects an equal number of Control and Treatment lanes in a single ROI (e.g. 3 Ctrl + 3 Treat).<br><br>"
            "<b>Lock ROI Size:</b> Forces any new ROI selection to have the exact same width and height as your previous selection. "
            "Essential for accurate Ponceau normalization across multiple treatments.<br><br>"
            "<b>Welch's Test:</b> If enabled, automatically performs a Welch's T-test between Control and Treatment groups in the summary."
        )
        QMessageBox.information(self, "Settings Help", help_msg)

    def show_results_help(self):
        help_msg = (
            "<h3>Results & Analysis Guide</h3>"
            "<b>Validation (SHOW PROFILE):</b><br>"
            "Click the blue <i>SHOW PROFILE</i> button to see the vertical intensity profile of your ROI. "
            "Red dashed lines indicate the lane separators. Use this to ensure the lane-specific background subtraction "
            "is accurate and not clipping signal peaks.<br><br>"
            "<b>Analysis Summary:</b><br>"
            "Shows the calculated averages for each group and the percentage change between Control and Treatment. "
            "If enabled, it also displays the P-value from a Welch's T-test.<br><br>"
            "<b>Detailed Data Points:</b><br>"
            "A breakdown of every lane analyzed. Columns show the raw intensities for Loading Control and Target protein, "
            "as well as the final normalized ratio.<br><br>"
            "<b>Excluding Replicates:</b><br>"
            "Right-click any row in the detailed table to <b>Exclude from Analysis</b>. "
            "Excluded samples are greyed out and automatically removed from Summary averages and P-value calculations."
        )
        QMessageBox.information(self, "Results Help", help_msg)

    def show_about(self):
        about_dialog = QDialog(self)
        about_dialog.setWindowTitle("About BlotQuant")
        about_dialog.setMinimumSize(550, 650)
        
        layout = QVBoxLayout(about_dialog)
        
        # Title
        title_label = QLabel("BlotQuant")
        title_label.setStyleSheet("font-size: 26px; font-weight: bold; color: #2c3e50;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        tabs = QTabWidget()
        
        # Tab 1: Get Started
        general_tab = QWidget()
        gen_layout = QVBoxLayout(general_tab)
        
        start_text = QLabel(
            "<h3>Get Started</h3>"
            "<b>1. Load Image:</b> Import your Western Blot (TIF, PNG, JPG).<br><br>"
            "<b>2. Align Blot:</b> Use the <i>Rotation Alignment</i> slider to ensure bands are perfectly vertical.<br><br>"
            "<b>3. Setup Analysis:</b> Enter the <i>Protein Name</i> and choose the <i>Type</i> (Loading Control or Target) and <i>Group</i> (Control or Treatment).<br><br>"
            "<b>4. Group History:</b> The <i>Group Name</i> field is now a dropdown that saves your entries. Quickly re-select 'Ctrl' or 'Insulin 10nM' when analyzing multiple blots from the same experiment.<br><br>"
            "<b>5. Select ROI:</b> Draw a rectangle around your bands. BlotQuant will automatically place lane separators.<br><br>"
            "<b>6. Validate Profile:</b> Click <i>SHOW PROFILE</i> to see a vertical intensity plot. Ensure the background subtraction (per lane) isn't cutting off your signal peaks.<br><br>"
            "<b>7. Apply & View:</b> Click <i>APPLY SELECTION</i>. Results, normalization, and statistical summaries update instantly in the right panel."
        )
        start_text.setWordWrap(True)
        gen_layout.addWidget(start_text)
        gen_layout.addStretch()
        tabs.addTab(general_tab, "Get Started")
        
        # Tab 2: Methodology
        method_tab = QWidget()
        method_layout = QVBoxLayout(method_tab)
        
        method_text = QLabel(
            "<h3>Densitometry Method</h3>"
            "<b>Lane-Specific Background Subtraction:</b><br>"
            "Calculates the <i>Median Intensity</i> individually for each lane within its separator boundaries. This compensates for background gradients (e.g., darker on one side) and prevents artifacts from affecting distant bands.<br><br>"
            "<b>Signal Integration:</b><br>"
            "Quantifies signal only from pixels exceeding a dynamic threshold (Background + 0.2 * SD). This excludes noise while capturing the full band signal.<br><br>"
            "<b>Vertical Intensity Profile:</b><br>"
            "Use the <i>SHOW PROFILE</i> button to visualize mean pixel intensity across the ROI width. Red dashed lines indicate lane separators, allowing you to verify that background subtraction is accurate and not clipping signal peaks.<br><br>"
            "<b>Normalization:</b><br>"
            "Automatically divides Target Protein intensities by their corresponding Loading Control (e.g., Actin) values from the same lane."
        )
        method_text.setWordWrap(True)
        method_layout.addWidget(method_text)
        method_layout.addStretch()
        tabs.addTab(method_tab, "Methodology")
        
        # Tab 3: Credits
        credits_tab = QWidget()
        cred_layout = QVBoxLayout(credits_tab)
        
        info_layout = QFormLayout()
        info_layout.addRow("Version:", QLabel(self.version))
        info_layout.addRow("Created:", QLabel(self.creation_date))
        info_layout.addRow("Author:", QLabel(self.author))
        
        affiliation_label = QLabel(self.affiliation)
        affiliation_label.setWordWrap(True)
        info_layout.addRow("Affiliation:", affiliation_label)
        
        website_label = QLabel('<a href="https://www.uni-potsdam.de/de/mem/index">https://www.uni-potsdam.de/de/mem/index</a>')
        website_label.setOpenExternalLinks(True)
        info_layout.addRow("Website:", website_label)
        
        # License Info
        license_label = QLabel("Intended for academic use under the MIT License.")
        license_label.setWordWrap(True)
        license_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        info_layout.addRow("License:", license_label)
        
        cred_layout.addLayout(info_layout)
        
        desc_label = QLabel("\nAutomated Western Blot densitometry for\nprecise protein quantification and normalization.")
        desc_label.setAlignment(Qt.AlignCenter)
        desc_label.setStyleSheet("font-style: italic; color: #7f8c8d;")
        cred_layout.addWidget(desc_label)
        
        cred_layout.addStretch()
        tabs.addTab(credits_tab, "Credits")
        
        layout.addWidget(tabs)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(about_dialog.accept)
        layout.addWidget(close_btn)
        
        about_dialog.exec()

    def load_image(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Load Image", "", "Image files (*.png *.jpg *.tif *.tiff)")
        if file_path:
            self.image = cv2.imread(file_path)
            self.rotation_slider.setValue(0) # Reset rotation on new image
            self.rotation_angle = 0
            self.display_image()

    def rotate_image(self, value):
        self.rotation_angle = value
        self.display_image()

    def reset_rotation(self):
        self.rotation_slider.setValue(0)
        self.rotation_angle = 0
        self.display_image()

    def display_image(self):
        if self.image is None:
            return
            
        # Convert BGR to RGB
        image_rgb = cv2.cvtColor(self.image, cv2.COLOR_BGR2RGB)
        h, w, ch = image_rgb.shape
        
        # Rotate if needed
        if self.rotation_angle != 0:
            # We'll use PIL for high-quality rotation matching the original logic
            image_pil = Image.fromarray(image_rgb)
            image_pil = image_pil.rotate(self.rotation_angle, Image.BICUBIC, expand=True)
            image_rgb = np.array(image_pil)
            h, w, ch = image_rgb.shape
            
        # Scale for display
        display_size = (800, 600)
        # Note: In PySide6 we can scale the QPixmap, but to match logic exactly
        # we might want to scale the array/PIL image first
        qimg = QImage(image_rgb.data, w, h, ch * w, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(qimg)
        
        scaled_pixmap = pixmap.scaled(800, 600, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.display_width = scaled_pixmap.width()
        self.display_height = scaled_pixmap.height()
        
        self.scene.clear()
        self.graphics_view.clear_selection()
        self.current_roi_rect = None
        self.separators = []
        
        self.scene.addPixmap(scaled_pixmap)
        self.scene.setSceneRect(0, 0, self.display_width, self.display_height)

    def add_protein(self):
        name = self.prot_name_input.text().strip()
        if name and name not in self.proteins:
            self.proteins[name] = {'regions': []}
            self.prot_combo.addItem(name)
            self.prot_combo.setCurrentText(name)
            self.current_protein = name
            self.prot_name_input.clear()

    def on_protein_selected(self, index):
        self.current_protein = self.prot_combo.currentText()

    def on_roi_selected(self, rect):
        # ROI Dimension Locking Logic
        if self.lock_roi_check.isChecked() and hasattr(self, 'last_roi_rect') and self.last_roi_rect:
            # Maintain width and height of the last ROI, but use current top-left
            rect = QRectF(rect.topLeft().x(), rect.topLeft().y(), 
                          self.last_roi_rect.width(), self.last_roi_rect.height())
            # Update the rectangle in the view to match the locked size
            self.graphics_view.current_rect_item.setRect(rect)

        self.current_roi_rect = rect
        # Only create new separators if the ROI size/position actually changed significantly
        # or if we don't have any yet.
        # This prevents snapping back while dragging separators.
        
        # NEW LOGIC: If we are actively DRAWING (first drag), always regenerate for even spacing.
        # If we are NOT drawing (e.g., resizing later or dragging separators), be more conservative.
        is_initial_drawing = self.graphics_view.is_drawing
        
        if not self.graphics_view.is_dragging_sep:
            if not self.separators or is_initial_drawing or self._should_regenerate_seps(rect):
                self.update_separators()
            else:
                # Just update separator heights to match the new ROI height
                # AND maintain their RELATIVE horizontal positions within the new width
                for sep in self.separators:
                    line = sep.line()
                    # Keep X position absolute - do not recalculate based on percentage
                    # to allow manual dragging to persist during ROI height/Y adjustments
                    line.setP1(QPointF(line.p1().x(), rect.top()))
                    line.setP2(QPointF(line.p2().x(), rect.bottom()))
                    sep.setLine(line)
        self.last_roi_rect = rect

    def _should_regenerate_seps(self, new_rect):
        """Helper to decide if we should reset separator positions."""
        if not self.separators: return True
        
        # If the number of expected separators would change, regenerate
        replicate_count = self.reps_spin.value()
        equal_samples = self.equal_n_check.isChecked()
        expected_num = (replicate_count * 2) - 1 if equal_samples else replicate_count - 1
        if len(self.separators) != expected_num:
            return True

        # If width changed significantly (>5%), regenerate
        # We use a smaller threshold than 10% to be more responsive to intentional resizing
        if hasattr(self, 'last_roi_rect'):
            if abs(new_rect.width() - self.last_roi_rect.width()) > (new_rect.width() * 0.05):
                return True
        return False

    def update_separators(self):
        # Clear old from both scene and tracking list
        for sep in self.separators:
            self.scene.removeItem(sep)
        self.separators = []
        
        if not self.graphics_view.current_rect_item:
            return
            
        rect = self.graphics_view.current_rect_item.rect()
        width = rect.width()
        x1 = rect.left()
        y1 = rect.top()
        y2 = rect.bottom()
        
        replicate_count = self.reps_spin.value()
        equal_samples = self.equal_n_check.isChecked()
        
        if equal_samples:
            num_seps = (replicate_count * 2) - 1
        else:
            num_seps = replicate_count - 1

        for i in range(num_seps):
            x_pos = x1 + width * (i + 1) / (num_seps + 1)
            sep = QGraphicsLineItem(x_pos, y1, x_pos, y2)
            sep.setPen(self.graphics_view.sep_pen)
            self.scene.addItem(sep)
            self.separators.append(sep)
        
        # Sync the AnalysisCanvas list
        self.graphics_view.separators = self.separators

    def apply_selection(self):
        # Prefer the actual item from the view, but fallback to current_roi_rect if needed
        rect_item = self.graphics_view.current_rect_item
        
        if rect_item or self.current_roi_rect:
            # Check if name is provided
            protein_name = self.protein_name_input.text().strip()
            if not protein_name:
                QMessageBox.warning(self, "Error", "Please enter a Protein/Blot Name.")
                return
            
            # Save group detail to history if it's new
            group_detail = self.treatment_detail_input.currentText().strip()
            if group_detail and group_detail not in self.group_history:
                self.group_history.append(group_detail)
                self.treatment_detail_input.addItem(group_detail)
            
            # Get ROI
            if rect_item:
                rect = rect_item.rect()
            else:
                rect = self.current_roi_rect
                
            roi_tuple = (rect.x(), rect.y(), rect.width(), rect.height())
            
            # Get manual separator positions (x-coordinates)
            sep_positions = [sep.line().p1().x() for sep in self.separators]
            
            self.process_roi(roi_tuple, sep_positions)
            self.graphics_view.clear_selection()
            self.current_roi_rect = None
            self.separators = []

    def extract_roi_data(self, roi, sep_positions=None):
        canvas_x, canvas_y, canvas_w, canvas_h = roi
        if canvas_w <= 0 or canvas_h <= 0 or self.image is None:
            return None, None

        # 1. Extraction (Rotating if needed)
        if self.rotation_angle != 0:
            height, width = self.image.shape[:2]
            center = (width // 2, height // 2)
            matrix = cv2.getRotationMatrix2D(center, self.rotation_angle, 1.0)
            
            cos = np.abs(matrix[0, 0])
            sin = np.abs(matrix[0, 1])
            new_w = int((height * sin) + (width * cos))
            new_h = int((height * cos) + (width * sin))
            matrix[0, 2] += (new_w / 2) - center[0]
            matrix[1, 2] += (new_h / 2) - center[1]
            
            rotated_full = cv2.warpAffine(self.image, matrix, (new_w, new_h), 
                                        flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
            
            scale_x = rotated_full.shape[1] / self.display_width
            scale_y = rotated_full.shape[0] / self.display_height
            
            x_adj = int(canvas_x * scale_x)
            y_adj = int(canvas_y * scale_y)
            w_adj = int(canvas_w * scale_x)
            h_adj = int(canvas_h * scale_y)
            
            roi_image = rotated_full[y_adj:y_adj+h_adj, x_adj:x_adj+w_adj]
            
            if sep_positions:
                scaled_seps = [int((sx - canvas_x) * scale_x) for sx in sep_positions]
            else:
                scaled_seps = None
        else:
            scale_x = self.image.shape[1] / self.display_width
            scale_y = self.image.shape[0] / self.display_height
            x_adj = int(canvas_x * scale_x)
            y_adj = int(canvas_y * scale_y)
            w_adj = int(canvas_w * scale_x)
            h_adj = int(canvas_h * scale_y)
            roi_image = self.image[y_adj:y_adj+h_adj, x_adj:x_adj+w_adj]
            
            if sep_positions:
                scaled_seps = [int((sx - canvas_x) * scale_x) for sx in sep_positions]
            else:
                scaled_seps = None

        if roi_image.size == 0:
            return None, None
            
        gray_roi = cv2.cvtColor(roi_image, cv2.COLOR_BGR2GRAY)
        # Ensure darker bands have HIGHER numerical values for intensity calculation
        gray_roi = cv2.bitwise_not(gray_roi)
        
        return gray_roi, scaled_seps

    def show_profile(self):
        rect_item = self.graphics_view.current_rect_item
        if not rect_item or self.image is None:
            QMessageBox.warning(self, "Error", "Please select an ROI first.")
            return
            
        rect = rect_item.rect()
        roi_tuple = (rect.x(), rect.y(), rect.width(), rect.height())
        sep_positions = [sep.line().p1().x() for sep in self.separators]
        
        gray_roi, scaled_seps = self.extract_roi_data(roi_tuple, sep_positions)
        if gray_roi is not None:
            dialog = ProfileDialog(gray_roi, scaled_seps or [], self)
            dialog.exec()

    def process_roi(self, roi, sep_positions=None):
        gray_roi, scaled_seps = self.extract_roi_data(roi, sep_positions)
        if gray_roi is None:
            return
        
        # 3. Lane Separation Logic with Lane-Specific Background
        intensities = []
        
        # Determine boundaries
        if scaled_seps:
            sorted_seps = sorted(scaled_seps)
            boundaries = [0] + sorted_seps + [gray_roi.shape[1]]
        else:
            replicate_count = self.reps_spin.value()
            equal_samples = self.equal_n_check.isChecked()
            num_sections = (replicate_count * 2) if equal_samples else replicate_count
            section_width = gray_roi.shape[1] / num_sections
            boundaries = [int(i * section_width) for i in range(num_sections + 1)]

        for i in range(len(boundaries) - 1):
            s_x = max(0, boundaries[i])
            e_x = min(gray_roi.shape[1], boundaries[i+1])
            
            if e_x <= s_x:
                intensities.append(0.0)
                continue
                
            section = gray_roi[:, s_x:e_x]
            
            # Lane-specific background (median of this lane only)
            lane_background = np.median(section)
            # More sensitive thresholding per lane
            lane_threshold = lane_background + (np.std(section) * 0.2)
            _, lane_mask = cv2.threshold(section, lane_threshold, 255, cv2.THRESH_BINARY)
            
            signal = section.astype(np.float32) - lane_background
            signal[signal < 0] = 0
            band_signal = signal[lane_mask > 0]
            
            if len(band_signal) > 0:
                intensities.append(np.sum(band_signal))
            else:
                intensities.append(0.0)

        # 4. Storage and Display
        group_type = "Treatment" if self.group_treat.isChecked() else "Control"
        mode = "Loading Control" if self.mode_control.isChecked() else "Target"
        protein_name = self.protein_name_input.text().strip() or mode
        treatment_detail = self.treatment_detail_input.currentText().strip() or ("None" if group_type == "Control" else "Generic Treatment")
        equal_samples = self.equal_n_check.isChecked()
        
        # Apply Start Index Offset
        start_offset = self.start_idx_spin.value() - 1  # 0-indexed offset
        if start_offset > 0:
            intensities = [0.0] * start_offset + intensities
        
        if equal_samples:
            mid = len(intensities) // 2
            ctrl_ints = intensities[:mid]
            treat_ints = intensities[mid:]
            
            # Save Control Part
            self.analysis_history.append({
                'type': mode,
                'group': 'Control',
                'detail': 'None',
                'name': protein_name,
                'intensities': ctrl_ints
            })
            
            # Save Treatment Part
            self.analysis_history.append({
                'type': mode,
                'group': 'Treatment',
                'detail': treatment_detail,
                'name': protein_name,
                'intensities': treat_ints
            })
        else:
            self.analysis_history.append({
                'type': mode,
                'group': group_type,
                'detail': treatment_detail,
                'name': protein_name,
                'intensities': intensities
            })
        
        self.refresh_analysis()

    def refresh_analysis(self):
        self.results_tree.clear()
        
        # 1. Gather all data
        lcs = [e for e in self.analysis_history if e['type'] == 'Loading Control']
        targets = [e for e in self.analysis_history if e['type'] == 'Target']
        
        if not lcs and not targets:
            return

        # 2. Define unique target names in order of analysis
        unique_target_names = []
        for t in targets:
            if t['name'] not in unique_target_names:
                unique_target_names.append(t['name'])

        # 3. Build dynamic headers
        # We'll show: Sample ID | [Active LC] | Target 1 (Raw) | Target 1 (Ratio) | Target 2 (Raw) | Target 2 (Ratio) ...
        # Note: "Active LC" column will show the LC used for the MOST RECENT target in that group for simplicity
        headers = ["Sample ID", "Loading Control"]
        for t_name in unique_target_names:
            headers.append(f"{t_name} (Raw)")
            headers.append(f"{t_name} (Ratio)")
        
        self.results_tree.setColumnCount(len(headers))
        self.results_tree.setHeaderLabels(headers)

        # 4. Process each group (Control, Treatment, etc.)
        all_groups = sorted(list(set(e['group'] for e in self.analysis_history)))
        
        # For the UI summary, we only want the LATEST target's result
        latest_normalized_data = {} # group -> list of normalized values for the latest target
        
        for group in all_groups:
            group_lcs = [e for e in lcs if e['group'] == group]
            group_targets = [e for e in targets if e['group'] == group]
            
            if not group_lcs and not group_targets:
                continue
                
            # Create group parent item (Control, Treatment, etc.)
            group_item = QTreeWidgetItem(self.results_tree)
            group_item.setText(0, group)
            group_item.setFont(0, QFont("", -1, QFont.Bold))
            group_item.setBackground(0, QColor("#f0f0f0"))
            
            # Find max replicates across all data for this group
            max_reps = 0
            for e in group_lcs + group_targets:
                max_reps = max(max_reps, len(e['intensities']))
            
            for i in range(max_reps):
                item = QTreeWidgetItem(group_item)
                is_excluded = i in self.excluded_samples.get(group, [])

                # Sample ID
                # Try to find a detail from the latest target, or LC
                detail = "Sample"
                if group_targets: detail = group_targets[-1]['detail']
                elif group_lcs: detail = group_lcs[-1]['detail']
                
                id_text = f"{detail} {i+1}" if detail not in ["None", "Generic Treatment"] else f"{group} {i+1}"
                if is_excluded: id_text += " [EXCLUDED]"
                item.setText(0, id_text)
                
                # Loading Control (show the most recent one for the group)
                lc_val = 0
                if group_lcs:
                    lc_entry = group_lcs[-1]
                    if i < len(lc_entry['intensities']):
                        lc_val = lc_entry['intensities'][i]
                        item.setText(1, f"{lc_val:.2f}")
                    else:
                        item.setText(1, "-")
                else:
                    item.setText(1, "-")
                
                # Target Columns
                col_idx = 2
                for t_name in unique_target_names:
                    # Find the specific target entry for this group and name
                    t_entries = [t for t in group_targets if t['name'] == t_name]
                    if t_entries:
                        t_entry = t_entries[-1]
                        
                        # Find the LC that was active when THIS target was analyzed
                        target_in_history_idx = self.analysis_history.index(t_entry)
                        applicable_lcs = [e for idx, e in enumerate(self.analysis_history) 
                                         if idx < target_in_history_idx and e['type'] == 'Loading Control' and e['group'] == group]
                        t_lc_entry = applicable_lcs[-1] if applicable_lcs else (group_lcs[-1] if group_lcs else None)
                        
                        t_val = t_entry['intensities'][i] if i < len(t_entry['intensities']) else 0
                        t_lc_val = t_lc_entry['intensities'][i] if t_lc_entry and i < len(t_lc_entry['intensities']) else 0
                        
                        # Raw Target
                        if t_val > 0:
                            item.setText(col_idx, f"{t_val:.2f}")
                        else:
                            item.setText(col_idx, "-")
                        
                        # Ratio
                        if t_lc_val > 0 and t_val > 0:
                            norm = t_val / t_lc_val
                            item.setText(col_idx + 1, f"{norm:.4f}")
                            
                            # For summary, if this is the latest target overall
                            if t_name == unique_target_names[-1] and not is_excluded:
                                if group not in latest_normalized_data: latest_normalized_data[group] = []
                                latest_normalized_data[group].append(norm)
                        else:
                            item.setText(col_idx + 1, "-")
                    else:
                        item.setText(col_idx, "-")
                        item.setText(col_idx + 1, "-")
                    
                    col_idx += 2

                # Exclusion styling
                if is_excluded:
                    for c in range(self.results_tree.columnCount()):
                        item.setForeground(c, QColor("#7f8c8d"))
                        font = item.font(c)
                        font.setStrikeOut(True)
                        item.setFont(c, font)
            
            self.results_tree.expandItem(group_item)

        # 5. Update Summary Display (Latest target only)
        if not latest_normalized_data:
            self.summary_text.setText("Analyze both Loading Control and Target to see results.")
            return

        summary_lines = []
        latest_target_name = unique_target_names[-1] if unique_target_names else "Target"
        summary_lines.append(f"<b>Latest Results ({latest_target_name}):</b>")
        
        group_means = {g: np.mean(v) for g, v in latest_normalized_data.items()}
        for g, m in group_means.items():
            summary_lines.append(f"&nbsp;&nbsp;{g} Average: {m:.4f}")
        
        groups = list(latest_normalized_data.keys())
        if "Control" in groups and "Treatment" in groups:
            ctrl_mean = group_means["Control"]
            treat_mean = group_means["Treatment"]
            if ctrl_mean != 0:
                pct_change = ((treat_mean - ctrl_mean) / ctrl_mean) * 100
                summary_lines.append(f"&nbsp;&nbsp;<b>Change:</b> {pct_change:+.2f}%")
        
        if self.ttest_check.isChecked() and "Control" in groups and "Treatment" in groups:
            ctrl_vals = latest_normalized_data["Control"]
            treat_vals = latest_normalized_data["Treatment"]
            if len(ctrl_vals) > 1 and len(treat_vals) > 1:
                _, p_val = stats.ttest_ind(ctrl_vals, treat_vals, equal_var=False)
                summary_lines.append(f"&nbsp;&nbsp;<b>P-Value (Welch's):</b> {p_val:.4f}")

        self.summary_text.setText("<br>".join(summary_lines))
        
        # Resize columns to contents for better readability
        for i in range(self.results_tree.columnCount()):
            self.results_tree.resizeColumnToContents(i)

    def analyze_blot(self):
        """Redundant: analysis now happens automatically in refresh_analysis."""
        self.refresh_analysis()

    def export_data(self):
        if not self.analysis_history:
            QMessageBox.warning(self, "Error", "No data to export.")
            return
            
        file_path, _ = QFileDialog.getSaveFileName(self, "Export Results", "", "Excel Files (*.xlsx)")
        if not file_path:
            return
            
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "BlotQuant_Results"
            
            # Global Font: Calibri
            calibri_font = "Calibri"
            
            # Styling formats
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Helper to apply border and font to a cell
            def style_cell(cell, bold=False, italic=False, underline=None, size=11, fill=None):
                cell.font = Font(name=calibri_font, size=size, bold=bold, italic=italic, underline=underline)
                cell.border = thin_border
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Helper to adjust column width
            def adjust_column_widths(worksheet):
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter # Get the column name
                    for cell in col:
                        try: # Necessary to avoid error on empty cells
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width

            # 1. Header
            ws['A1'] = "Western Blot Analysis Report"
            ws['A1'].font = Font(name=calibri_font, size=14, bold=True)
            ws['A2'] = f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ws['A2'].font = Font(name=calibri_font, size=11)
            
            exp_info = self.experiment_input.text().strip()
            if exp_info:
                ws['A3'] = f"Experiment: {exp_info}"
                ws['A3'].font = Font(name=calibri_font, size=11)
            
            # 2. Analysis Summary Section
            ws['A5'] = "ANALYSIS SUMMARY"
            ws['A5'].font = Font(name=calibri_font, size=12, bold=True, underline="single")
            
            lcs = [e for e in self.analysis_history if e['type'] == 'Loading Control']
            targets = [e for e in self.analysis_history if e['type'] == 'Target']
            
            # Calculate summary stats for the summary table
            normalized_results = {} # {target_name: {group: [vals]}}
            for target in targets:
                t_name = target['name']
                if t_name not in normalized_results: normalized_results[t_name] = {}
                
                # Find the most recent LC for this group that was defined BEFORE this target
                target_in_history_idx = self.analysis_history.index(target)
                applicable_lcs = [e for i, e in enumerate(self.analysis_history) 
                                 if i < target_in_history_idx and e['type'] == 'Loading Control' and e['group'] == target['group']]
                
                if applicable_lcs:
                    lc = applicable_lcs[-1]
                else:
                    matching_lcs = [lc for lc in lcs if lc['group'] == target['group']]
                    lc = matching_lcs[-1] if matching_lcs else None

                if lc:
                    lc_ints = lc['intensities']
                    t_ints = target['intensities']
                    min_len = min(len(lc_ints), len(t_ints))
                    norm_vals = [t_ints[i] / lc_ints[i] if lc_ints[i] != 0 else 0 for i in range(min_len)]
                    final_vals = [v for i, v in enumerate(norm_vals) if i not in self.excluded_samples.get(target['group'], [])]
                    normalized_results[t_name][target['group']] = final_vals

            curr_row = 7
            summary_headers = ["Target Protein", "Group", "Detail", "Mean", "SEM", "N", "Excluded Indices"]
            for col, text in enumerate(summary_headers, 1):
                cell = ws.cell(row=curr_row, column=col, value=text)
                style_cell(cell, bold=True, fill=header_fill)
            curr_row += 1

            for target_name, groups in normalized_results.items():
                for group_name, vals in groups.items():
                    detail = next((t['detail'] for t in targets if t['name'] == target_name and t['group'] == group_name), "")
                    
                    style_cell(ws.cell(row=curr_row, column=1, value=target_name))
                    style_cell(ws.cell(row=curr_row, column=2, value=group_name))
                    style_cell(ws.cell(row=curr_row, column=3, value=detail))
                    
                    if vals:
                        style_cell(ws.cell(row=curr_row, column=4, value=np.mean(vals)))
                        style_cell(ws.cell(row=curr_row, column=5, value=stats.sem(vals) if len(vals) > 1 else 0))
                        style_cell(ws.cell(row=curr_row, column=6, value=len(vals)))
                    else:
                        style_cell(ws.cell(row=curr_row, column=4, value="N/A"))
                        style_cell(ws.cell(row=curr_row, column=5, value="-"))
                        style_cell(ws.cell(row=curr_row, column=6, value="0"))
                    
                    excluded = self.excluded_samples.get(group_name, [])
                    if excluded:
                        style_cell(ws.cell(row=curr_row, column=7, value=", ".join(str(x+1) for x in excluded)))
                    else:
                        style_cell(ws.cell(row=curr_row, column=7, value="-"))
                    curr_row += 1
                
                if "Control" in groups and "Treatment" in groups:
                    c_vals = groups["Control"]
                    t_vals = groups["Treatment"]
                    
                    # Percent Change
                    if c_vals:
                        c_mean = np.mean(c_vals)
                        t_mean = np.mean(t_vals) if t_vals else 0
                        if c_mean != 0:
                            pct_change = ((t_mean - c_mean) / c_mean) * 100
                            cell1 = ws.cell(row=curr_row, column=1, value=f"% Change ({target_name} Ctrl vs Treat):")
                            style_cell(cell1, italic=True)
                            cell1.alignment = Alignment(horizontal='right')
                            
                            cell2 = ws.cell(row=curr_row, column=2, value=f"{pct_change:+.2f}%")
                            style_cell(cell2)
                            curr_row += 1

                    # T-Test
                    if len(c_vals) > 1 and len(t_vals) > 1:
                        _, p_val = stats.ttest_ind(c_vals, t_vals, equal_var=False)
                        cell1 = ws.cell(row=curr_row, column=1, value=f"P-value ({target_name} Ctrl vs Treat):")
                        style_cell(cell1, italic=True)
                        cell1.alignment = Alignment(horizontal='right')
                        
                        cell2 = ws.cell(row=curr_row, column=2, value=p_val)
                        style_cell(cell2)
                        curr_row += 1
                curr_row += 1 # Spacer

            # 3. Detailed Data Section
            curr_row += 1
            ws.cell(row=curr_row, column=1, value="DETAILED DATA (ALL TARGETS)").font = Font(name=calibri_font, size=12, bold=True, underline="single")
            curr_row += 2
            
            unique_target_names = []
            for t in targets:
                if t['name'] not in unique_target_names:
                    unique_target_names.append(t['name'])

            # Headers
            headers = ["Group", "Sample ID", "Loading Control"]
            for t_name in unique_target_names:
                headers.append(f"{t_name} (Raw)")
                headers.append(f"{t_name} (Ratio)")
            
            for col, text in enumerate(headers, 1):
                cell = ws.cell(row=curr_row, column=col, value=text)
                style_cell(cell, bold=True, fill=header_fill)
            curr_row += 1

            all_groups = sorted(list(set(e['group'] for e in self.analysis_history)))
            for group in all_groups:
                group_lcs = [e for e in lcs if e['group'] == group]
                group_targets = [e for e in targets if e['group'] == group]
                
                max_reps = 0
                for e in group_lcs + group_targets:
                    max_reps = max(max_reps, len(e['intensities']))
                
                for i in range(max_reps):
                    is_excluded = i in self.excluded_samples.get(group, [])
                    
                    # Group & ID
                    detail = "Sample"
                    if group_targets: detail = group_targets[-1]['detail']
                    elif group_lcs: detail = group_lcs[-1]['detail']
                    id_text = f"{detail} {i+1}" if detail not in ["None", "Generic Treatment"] else f"{group} {i+1}"
                    if is_excluded: id_text += " [EXCLUDED]"
                    
                    cell_group = ws.cell(row=curr_row, column=1, value=group)
                    cell_id = ws.cell(row=curr_row, column=2, value=id_text)
                    style_cell(cell_group)
                    style_cell(cell_id)
                    
                    # LC
                    lc_val = None
                    if group_lcs:
                        lc_entry = group_lcs[-1]
                        lc_val = lc_entry['intensities'][i] if i < len(lc_entry['intensities']) else None
                        cell_lc = ws.cell(row=curr_row, column=3, value=lc_val)
                        style_cell(cell_lc)
                    else:
                        style_cell(ws.cell(row=curr_row, column=3, value="-"))
                    
                    # Targets
                    col_offset = 4
                    for t_name in unique_target_names:
                        t_entries = [t for t in group_targets if t['name'] == t_name]
                        if t_entries:
                            t_entry = t_entries[-1]
                            target_in_history_idx = self.analysis_history.index(t_entry)
                            applicable_lcs = [e for idx, e in enumerate(self.analysis_history) 
                                             if idx < target_in_history_idx and e['type'] == 'Loading Control' and e['group'] == group]
                            t_lc_entry = applicable_lcs[-1] if applicable_lcs else (group_lcs[-1] if group_lcs else None)
                            
                            t_val = t_entry['intensities'][i] if i < len(t_entry['intensities']) else None
                            t_lc_val = t_lc_entry['intensities'][i] if t_lc_entry and i < len(t_lc_entry['intensities']) else None
                            
                            cell_raw = ws.cell(row=curr_row, column=col_offset, value=t_val)
                            style_cell(cell_raw)
                            
                            if t_val is not None and t_lc_val is not None and t_lc_val != 0:
                                cell_ratio = ws.cell(row=curr_row, column=col_offset + 1, value=t_val / t_lc_val)
                            else:
                                cell_ratio = ws.cell(row=curr_row, column=col_offset + 1, value="-")
                            style_cell(cell_ratio)
                        else:
                            style_cell(ws.cell(row=curr_row, column=col_offset, value="-"))
                            style_cell(ws.cell(row=curr_row, column=col_offset + 1, value="-"))
                        col_offset += 2
                    
                    # Style excluded rows
                    if is_excluded:
                        for col in range(1, col_offset):
                            cell = ws.cell(row=curr_row, column=col)
                            cell.font = Font(name=calibri_font, strike=True, color="7f8c8d")
                    
                    curr_row += 1
                curr_row += 1 # Spacer between groups
            
            adjust_column_widths(ws)
            wb.save(file_path)
            QMessageBox.information(self, "Success", f"Data exported to {file_path}")
            import os
            os.startfile(file_path)
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export: {str(e)}")

    def open_graphpad(self):
        import subprocess
        import os
        
        # Common installation paths for GraphPad Prism
        prism_paths = [
            r"C:\Program Files\GraphPad\Prism 10\Prism.exe",
            r"C:\Program Files\GraphPad\Prism 9\Prism.exe",
            r"C:\Program Files\GraphPad\Prism 8\Prism.exe",
            r"C:\Program Files\GraphPad\Prism 7\Prism.exe"
        ]
        
        found = False
        for path in prism_paths:
            if os.path.exists(path):
                try:
                    subprocess.Popen([path])
                    found = True
                    break
                except:
                    continue
        
        if not found:
            QMessageBox.warning(self, "Warning", 
                "GraphPad Prism could not be found in standard locations.\n"
                "Please open Prism manually.")

    def undo_last(self):
        if self.analysis_history:
            self.analysis_history.pop()
            self.refresh_analysis()

    def copy_to_clipboard(self):
        if not self.analysis_history:
            QMessageBox.warning(self, "Error", "No data to copy.")
            return

        # 1. Gather all data
        lcs = [e for e in self.analysis_history if e['type'] == 'Loading Control']
        targets = [e for e in self.analysis_history if e['type'] == 'Target']
        
        unique_target_names = []
        for t in targets:
            if t['name'] not in unique_target_names:
                unique_target_names.append(t['name'])

        # --- A. Prepare Data ---
        exp_info = self.experiment_input.text().strip()
        date_str = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        # Calculate summary stats for each target
        normalized_results = {} # {target_name: {group: [vals]}}
        for target in targets:
            t_name = target['name']
            if t_name not in normalized_results: normalized_results[t_name] = {}
            
            target_in_history_idx = self.analysis_history.index(target)
            applicable_lcs = [e for i, e in enumerate(self.analysis_history) 
                             if i < target_in_history_idx and e['type'] == 'Loading Control' and e['group'] == target['group']]
            lc = applicable_lcs[-1] if applicable_lcs else ([lc for lc in lcs if lc['group'] == target['group']][-1] if [lc for lc in lcs if lc['group'] == target['group']] else None)

            if lc:
                lc_ints = lc['intensities']
                t_ints = target['intensities']
                min_len = min(len(lc_ints), len(t_ints))
                norm_vals = [t_ints[i] / lc_ints[i] if lc_ints[i] != 0 else 0 for i in range(min_len)]
                final_vals = [v for i, v in enumerate(norm_vals) if i not in self.excluded_samples.get(target['group'], [])]
                normalized_results[t_name][target['group']] = final_vals

        # --- B. Build HTML Version (for Bold Headers) ---
        html_parts = ["<html><body>"]
        html_parts.append("""
        <style>
            table { border-collapse: collapse; width: 100%; font-family: sans-serif; }
            th, td { border: 1px solid #ccc; padding: 6px 10px; text-align: left; white-space: nowrap; }
            th { background-color: #f2f2f2; font-weight: bold; }
            .excluded { text-decoration: line-through; color: #7f8c8d; }
            h3 { margin-bottom: 5px; color: #2c3e50; }
            p { margin: 2px 0; font-size: 0.9em; }
        </style>
        """)
        html_parts.append("<h3>ANALYSIS SUMMARY</h3>")
        if exp_info:
            html_parts.append(f"<p><b>Experiment:</b> {exp_info}</p>")
        html_parts.append(f"<p><b>Export Date:</b> {date_str}</p>")
        
        html_parts.append("<table>")
        html_parts.append("<tr><th>Target</th><th>Group</th><th>Mean</th><th>SEM</th><th>N</th></tr>")
        
        for t_name in unique_target_names:
            groups_data = normalized_results.get(t_name, {})
            for g_name in sorted(groups_data.keys()):
                vals = groups_data[g_name]
                if vals:
                    mean = np.mean(vals)
                    sem = stats.sem(vals) if len(vals) > 1 else 0
                    html_parts.append(f"<tr><td>{t_name}</td><td>{g_name}</td><td>{mean:.4f}</td><td>{sem:.4f}</td><td>{len(vals)}</td></tr>")
            
            if "Control" in groups_data and "Treatment" in groups_data:
                c_vals = groups_data["Control"]
                t_vals = groups_data["Treatment"]
                if c_vals:
                    c_mean = np.mean(c_vals)
                    t_mean = np.mean(t_vals) if t_vals else 0
                    if c_mean != 0:
                        pct_change = ((t_mean - c_mean) / c_mean) * 100
                        html_parts.append(f"<tr><td colspan='2'><b>% Change ({t_name}):</b></td><td colspan='3'>{pct_change:+.2f}%</td></tr>")
                if len(c_vals) > 1 and len(t_vals) > 1:
                    _, p_val = stats.ttest_ind(c_vals, t_vals, equal_var=False)
                    html_parts.append(f"<tr><td colspan='2'><b>P-value ({t_name}):</b></td><td colspan='3'>{p_val:.4f}</td></tr>")
        html_parts.append("</table><br>")

        html_parts.append("<h3>DETAILED DATA</h3>")
        html_parts.append("<table border='1' style='border-collapse: collapse;'>")
        
        detailed_headers = ["Group", "Sample ID", "Loading Control"]
        for t_name in unique_target_names:
            detailed_headers.append(f"{t_name} (Raw)")
            detailed_headers.append(f"{t_name} (Ratio)")
        
        html_parts.append("<tr style='background-color: #f2f2f2;'>")
        for h in detailed_headers:
            html_parts.append(f"<th>{h}</th>")
        html_parts.append("</tr>")
        
        all_groups = sorted(list(set(e['group'] for e in self.analysis_history)))
        for group in all_groups:
            group_lcs = [e for e in lcs if e['group'] == group]
            group_targets = [e for e in targets if e['group'] == group]
            max_reps = 0
            for e in group_lcs + group_targets:
                max_reps = max(max_reps, len(e['intensities']))
            
            for i in range(max_reps):
                is_excluded = i in self.excluded_samples.get(group, [])
                style = "style='color: #7f8c8d; text-decoration: line-through;'" if is_excluded else ""
                html_parts.append(f"<tr {style}>")
                
                detail = "Sample"
                if group_targets: detail = group_targets[-1]['detail']
                elif group_lcs: detail = group_lcs[-1]['detail']
                id_text = f"{detail} {i+1}" if detail not in ["None", "Generic Treatment"] else f"{group} {i+1}"
                if is_excluded: id_text += " [EXCLUDED]"
                
                html_parts.append(f"<td>{group}</td><td>{id_text}</td>")
                
                if group_lcs:
                    lc_entry = group_lcs[-1]
                    val = lc_entry['intensities'][i] if i < len(lc_entry['intensities']) else 0
                    html_parts.append(f"<td>{val:.2f}</td>" if val > 0 else "<td>-</td>")
                else:
                    html_parts.append("<td>-</td>")
                
                for t_name in unique_target_names:
                    t_entries = [t for t in group_targets if t['name'] == t_name]
                    if t_entries:
                        t_entry = t_entries[-1]
                        target_in_history_idx = self.analysis_history.index(t_entry)
                        applicable_lcs = [e for idx, e in enumerate(self.analysis_history) 
                                         if idx < target_in_history_idx and e['type'] == 'Loading Control' and e['group'] == group]
                        t_lc_entry = applicable_lcs[-1] if applicable_lcs else (group_lcs[-1] if group_lcs else None)
                        
                        t_val = t_entry['intensities'][i] if i < len(t_entry['intensities']) else 0
                        t_lc_val = t_lc_entry['intensities'][i] if t_lc_entry and i < len(t_lc_entry['intensities']) else 0
                        
                        html_parts.append(f"<td>{t_val:.2f}</td>" if t_val > 0 else "<td>-</td>")
                        if t_val > 0 and t_lc_val > 0:
                            html_parts.append(f"<td>{t_val/t_lc_val:.4f}</td>")
                        else:
                            html_parts.append("<td>-</td>")
                    else:
                        html_parts.append("<td>-</td><td>-</td>")
                html_parts.append("</tr>")
        html_parts.append("</table>")
        html_parts.append("</body></html>")

        # --- C. Build Plain Text Version (for compatibility) ---
        text_parts = ["ANALYSIS SUMMARY"]
        if exp_info: text_parts.append(f"Experiment:\t{exp_info}")
        text_parts.append(f"Export Date:\t{date_str}\n")
        
        text_parts.append("Target\tGroup\tMean\tSEM\tN")
        for t_name in unique_target_names:
            groups_data = normalized_results.get(t_name, {})
            for g_name in sorted(groups_data.keys()):
                vals = groups_data[g_name]
                if vals:
                    mean = np.mean(vals)
                    sem = stats.sem(vals) if len(vals) > 1 else 0
                    text_parts.append(f"{t_name}\t{g_name}\t{mean:.4f}\t{sem:.4f}\t{len(vals)}")
            if "Control" in groups_data and "Treatment" in groups_data:
                c_vals = groups_data["Control"]
                t_vals = groups_data["Treatment"]
                if c_vals:
                    c_mean = np.mean(c_vals)
                    t_mean = np.mean(t_vals) if t_vals else 0
                    if c_mean != 0:
                        pct_change = ((t_mean - c_mean) / c_mean) * 100
                        text_parts.append(f"% Change ({t_name}):\t{pct_change:+.2f}%")
                if len(c_vals) > 1 and len(t_vals) > 1:
                    _, p_val = stats.ttest_ind(c_vals, t_vals, equal_var=False)
                    text_parts.append(f"P-value ({t_name}):\t{p_val:.4f}")
            text_parts.append("")

        text_parts.append("DETAILED DATA")
        text_parts.append("\t".join(detailed_headers))
        for group in all_groups:
            group_lcs = [e for e in lcs if e['group'] == group]
            group_targets = [e for e in targets if e['group'] == group]
            max_reps = 0
            for e in group_lcs + group_targets:
                max_reps = max(max_reps, len(e['intensities']))
            for i in range(max_reps):
                row = []
                is_excluded = i in self.excluded_samples.get(group, [])
                detail = "Sample"
                if group_targets: detail = group_targets[-1]['detail']
                elif group_lcs: detail = group_lcs[-1]['detail']
                id_text = f"{detail} {i+1}" if detail not in ["None", "Generic Treatment"] else f"{group} {i+1}"
                if is_excluded: id_text += " [EXCLUDED]"
                row.extend([group, id_text])
                if group_lcs:
                    lc_entry = group_lcs[-1]
                    val = lc_entry['intensities'][i] if i < len(lc_entry['intensities']) else 0
                    row.append(f"{val:.2f}" if val > 0 else "-")
                else: row.append("-")
                for t_name in unique_target_names:
                    t_entries = [t for t in group_targets if t['name'] == t_name]
                    if t_entries:
                        t_entry = t_entries[-1]
                        target_in_history_idx = self.analysis_history.index(t_entry)
                        applicable_lcs = [e for idx, e in enumerate(self.analysis_history) 
                                         if idx < target_in_history_idx and e['type'] == 'Loading Control' and e['group'] == group]
                        t_lc_entry = applicable_lcs[-1] if applicable_lcs else (group_lcs[-1] if group_lcs else None)
                        t_val = t_entry['intensities'][i] if i < len(t_entry['intensities']) else 0
                        t_lc_val = t_lc_entry['intensities'][i] if t_lc_entry and i < len(t_lc_entry['intensities']) else 0
                        row.append(f"{t_val:.2f}" if t_val > 0 else "-")
                        row.append(f"{t_val/t_lc_val:.4f}" if t_val > 0 and t_lc_val > 0 else "-")
                    else: row.extend(["-", "-"])
                text_parts.append("\t".join(row))

        # --- D. Set Clipboard ---
        mime_data = QMimeData()
        mime_data.setText("\n".join(text_parts))
        mime_data.setHtml("\n".join(html_parts))
        
        QApplication.clipboard().setMimeData(mime_data)
        QMessageBox.information(self, "Success", "Results copied to clipboard (with bold headers for Excel/Word).")

    def start_over(self):
        reply = QMessageBox.question(self, "Confirm Reset", 
                                   "Are you sure you want to clear all data and start over?",
                                   QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            self.analysis_history = []
            self.excluded_samples = {}
            self.results_tree.clear()
            self.results_tree.setColumnCount(2)
            self.results_tree.setHeaderLabels(["Sample ID", "Loading Control"])
            self.scene.clear()
            self.image = None
            self.rotation_slider.setValue(0)
            self.graphics_view.clear_selection()
            self.current_roi_rect = None
            self.separators = []
            self.experiment_input.clear()
            self.summary_text.setText("No analysis performed yet.")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = BlotQuant()
    window.show()
    sys.exit(app.exec())
